import argparse
import jieba
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openai import AzureOpenAI
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from tool.identify_catrgory import check_category

# construct class to encapsulation data
class data_loader():
    # constructor
    def __init__(self):
        self.summarization_sentences, self.case_numbers = self.get_summarization()
        self.FAQ_Questions, self.FAQ_Categories, self.FAQ_Answers  = self.get_info()

    # 1. get the mass and worker summarization from file (第_批逐字稿摘要.xlsx)
    def get_summarization(self):
        # construct variable (file_name)
        file_path = opt.excel_name
        # construct list for stored mass sentences
        mass_sentences, mass_sentence_case_numbers = [], []
        # construct two list for stored case_numbers and mass_summarizations
        case_numbers, mass_summarizations = [], []
        # try-catch format
        try:
            # read file (第_批逐字稿摘要.xlsx)
            file = pd.read_excel(file_path, usecols=f'B:{opt.excel_range}', header=0)
            # loop for row ["case number", "mass summarization"]
            for row_num in [0, 49]:
                # construct variable (row) to get the summarization
                row = file.iloc[row_num]
                # drop out the unname and nan
                cleaned_row = row[~row.index.str.contains('Unnamed', na=False)]
                cleaned_row.dropna()

                if row_num == 0:
                    # construct list (extractive_summary) to store summary
                    case_numbers = cleaned_row.tolist()
                elif row_num == 49:
                    # drop out the space of the summarization
                    mass_summarizations = cleaned_row.tolist() #[summary for summary in cleaned_row.tolist() if summary != " "]

            # loop for list (mass_summarization)
            for index, summary in enumerate(mass_summarizations):
                if summary != " ":
                    # according the "\n" to split summary
                    sentences = summary.split("\n")
                    # stored it into list(mass_sentences)
                    mass_sentences.extend(sentences)
                    # stored the case_number into list (mass_case_numbers)
                    mass_sentence_case_numbers.extend([case_numbers[index]] * len(sentences))
                

            #vertify data
            print("-----------------")
            print(f"mass sentences: {mass_sentences[0]}, len: {len(mass_sentences)}")
            print("-----------------")
            print(f"mass case number: {mass_sentence_case_numbers[0]}, len: {len(mass_sentence_case_numbers)}")
            # return value
            return mass_sentences, mass_sentence_case_numbers
        
        except Exception as e:
            print(f"找不到目前的文件_{file_path}")
        except FileExistsError:
            print(f"您遇到的問題為_{e}")

    # 2. get info (問題) from file(裁決中心常見問題)
    def get_info(self):
        # construct variable (FAQ_path)
        FAQ_path = opt.FAQ
        # construct list (FAQ_Info) to record info 
        FAQ_Info, FAQ_categories, FAQ_answers = [], [], []
        # try-catch format
        try:
            # construct variable (wb) to represent wordbook
            wb = load_workbook(FAQ_path, data_only=True)
            # construct variable (FAQ) to get page ("常見問題20250429") form wb
            FAQ = wb["常見問題20250519"]

            # loop for paremeter (FAQ_range), start from 2
            for index in range(2, opt.FAQ_row+1):
                # construct variable (info) to record info
                FAQ_Info.append(FAQ.cell(index, 5).value)
                FAQ_categories.append(FAQ.cell(index, 4).value)
                FAQ_answers.append(FAQ.cell(index, 7).value)

            #vertify data
            print("-----------------")
            print(f"FAQ: {FAQ_Info[0]}, len: {len(FAQ_Info)}\n")
            # return value
            return FAQ_Info, FAQ_categories, FAQ_answers

        except Exception as e:
            print(f"您遇到的問題為_{e}")
        except FileExistsError:
            print(f"找不到目前的文件_{FAQ_path}")

# 3. get gpt result follwing summarization(should replace the space val by worker summarization) from first step and info (知識編號, 分類, 問題)
def get_gpt_result():
    # create new object of class (dataloader)
    data = data_loader()
    # construct list (gpt_result) to stored the answer
    gpt_result = []
    # variable count which seen as a number
    count = 0
    # construct a TF-IDF vector translater
    vectorizer = TfidfVectorizer(
        tokenizer=lambda x: jieba.cut(x, cut_all=False),
        token_pattern=None 
        # use personal tokenizer
    )

    # construct variant (end_point, deployment, subscription, api_version, client, example)
    endpoint = "https://aicodeassistant.openai.azure.com/"
    deployment = "gpt-4o-mini"

    subscription = ""
    api_version = "2024-12-01-preview"

    # follwing the above parameters to construct client
    client = AzureOpenAI(
        api_version=api_version,
        azure_endpoint=endpoint,
        api_key=subscription,
    )

    for indices in range(len(data.summarization_sentences)):
        # construct prompt to give gpt
        prompt = f"裁決中心常見問題: \n\"\"\"{data.FAQ_Questions}\"\"\"\n, 其中包含民眾常見的問題。\
                   萃取摘要句子: \n\"\"\"{data.summarization_sentences[indices]}\"\"\"\n, 為民眾闡述的句子。\
                   接下來要麻煩您根據提供的句子, 與裁決中心常見問題中的每一個問題進行比較，找出其中最相關的三個問題(一定要回傳三個)。\
                   生成範例:\n\"\"\" 結果: 申請裁決書可以委由他人代辦嗎？、我被攔停舉發怎麼申請裁決書？、我收到裁決書，怎麼沒有附照片？ \"\"\"\n\
                   生成規範:三種案件編號需要以(、)進行連結並在句子開頭加上(結果:)。另外不能更改常見問題中的文字。結果生成時不需要加上思考過程, 直接回傳即可。\
                  "
        
        # give the request (includes role、content) to gpt and get the respond from gpt
        response = client.chat.completions.create(
            messages=[{
                "role": "system",
                "content": "You are a helpful assistant."
            },{
                "role": "user",
                "content": prompt,
            }],
            max_tokens=4096,
            temperature=1.0,
            top_p=1.0,
            model=deployment
        )

        # loop for response
        if ":" in response.choices[0].message.content and response.choices[0].message.content != None:
            responses = response.choices[0].message.content.split(":")[1]
        else:
            # avoid the program breaking
            continue

        # because of AI hallucination, translate the result to the question which is most similar
        for response in responses.split("、"):
            # do additional action (similarity) to avoid the llm change the word in response
            all_content = data.FAQ_Questions + [response]
            # send all_text to TF-IDF vector transklater
            tfidf_matrix = vectorizer.fit_transform(all_content)
            # construct two variant (sentence_vectors, summary_vectors) to get sentence and summary vectors
            questions_vectors, response_vector = tfidf_matrix[:-1], tfidf_matrix[-1:]
            # caculate cosine similarity between sentence and summary
            sentence_similarities = cosine_similarity(questions_vectors, response_vector)
            #get top ext_len-th of sentences
            top_sentence_indices = np.argsort(sentence_similarities.flatten())[-1]
            selected_response = data.FAQ_Questions[top_sentence_indices]

            category, answer = check_category(selected_response, data.FAQ_Questions, data.FAQ_Categories, data.FAQ_Answers)

            # store the result to list (masses)
            gpt_result.append([str(count), data.case_numbers[indices], data.summarization_sentences[indices], selected_response, answer, category])
            # update variable (count)
            count += 1

        print(f"\r取出前三標註建議-執行進度:[{indices}/{len(data.summarization_sentences)-1}]", end="")

    # vertifi data
    print("\n-----------------")
    print(f"result: {gpt_result[0]}, len: {len(gpt_result)}")

    #call the follwing function
    update_file(gpt_result)

# 4. update the file (第_批逐字稿摘要.xlsx)
def update_file(gpt_results):
    # construct file path
    file_path = opt.FAQ_Result

    try:
        # construct variable (wb) to load the page
        wb = load_workbook(file_path, data_only=True)
        # construct variable (wb_page)
        wb_page = wb["工作表1"]
        # loop for paremeter (gpt_result)
        for result in gpt_results: wb_page.append(result)
        # save file
        wb.save(file_path)
        print("\n-----------------")
        print("結果已經更新至檔案中")

    except Exception as e:
        print(f"您遇到的問題為_{e}")

# python paremeter excecl_name, FAQ, range
if __name__ == "__main__":
    # construct variant (parser) to set argparser lib
    parser = argparse.ArgumentParser()
    # add argurment
    parser.add_argument("--excel_name", type=str, default="250221_逐字稿摘要.xlsx")
    parser.add_argument("--FAQ", type=str, default="裁決中心常見問題(標出標準問句)_250519.xlsx")
    parser.add_argument("--FAQ_Result", type=str, default="FAQ_Result.xlsx")
    parser.add_argument("--excel_range", type=str, default="AFY")
    parser.add_argument("--FAQ_row", type=int, default=251)
    # construct variant (opt) to present paremeter
    opt = parser.parse_args()
    print(opt)
    get_gpt_result()