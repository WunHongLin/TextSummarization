from openpyxl import load_workbook
from matplotlib import pyplot as plt
import matplotlib

matplotlib.rc('font', family='Microsoft JhengHei')

wb = load_workbook("裁決中心常見問題(標出標準問句)_250519.xlsx")
wp = wb["常見問題20250519"]

number = []

for index in range(2, 252):
    questions = wp.cell(index, 6).value.split("\n")
    number.append(len(questions))
    wp.cell(index, 10).value = len(questions)

print("已儲存...")
wb.save("裁決中心常見問題(標出標準問句)_250519.xlsx")

fre_dict = {k: 0 for k in range(1, max(number)+1)}
for item in number: fre_dict[item] += 1

fig = plt.figure(figsize=(15,8))
bar = plt.bar(list(range(len(number))), number)
# plt.bar_label(bar)
plt.title("示意圖")
plt.xlabel("知識編號")
plt.ylabel("題目數量")

plt.savefig("./result1.png",
            transparent=True,
            bbox_inches='tight',
            pad_inches=1)
plt.show()

#print(f"\r {index:03d}/ {len(doc_content)} [{"█"*(index//50)}{" "*((len(doc_content)-index)//50)}] {round(index*100/len(doc_content), 2)}%", end="")


